from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .exceptions import ValidationError
from .models import QuestionItem, StudentScores

REQUIRED_HEADERS = ["题号", "题型", "核心考点", "分值", "难度等级", "易错点"]
NAME_HEADER = "姓名"
TOTAL_HEADER = "全卷分数"


def _as_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _as_float(value: object, label: str) -> float:
    if value is None or value == "":
        raise ValidationError(f"{label}不能为空")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValidationError(f"{label}必须是数字，当前为：{value}") from exc


def _normal_header(value: object) -> str:
    return re.sub(r"\s+", "", _as_text(value))


def load_detail_table(path: str | Path) -> List[QuestionItem]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    header_map = {_normal_header(cell.value): idx for idx, cell in enumerate(sheet[1], start=1)}

    missing = [header for header in REQUIRED_HEADERS if header not in header_map]
    if missing:
        raise ValidationError(f"细目表缺少必需列：{', '.join(missing)}")

    items: List[QuestionItem] = []
    seen_numbers: set[str] = set()
    for row_index in range(2, sheet.max_row + 1):
        number = _as_text(sheet.cell(row_index, header_map["题号"]).value)
        if not number:
            continue
        if number in seen_numbers:
            raise ValidationError(f"细目表第 {row_index} 行题号重复：{number}")
        seen_numbers.add(number)

        question_type = _as_text(sheet.cell(row_index, header_map["题型"]).value)
        core_concept = _as_text(sheet.cell(row_index, header_map["核心考点"]).value)
        difficulty = _as_text(sheet.cell(row_index, header_map["难度等级"]).value)
        misconception = _as_text(sheet.cell(row_index, header_map["易错点"]).value)
        score = _as_float(sheet.cell(row_index, header_map["分值"]).value, f"第 {row_index} 行分值")

        if not question_type:
            raise ValidationError(f"细目表第 {row_index} 行题型不能为空")
        if not core_concept:
            raise ValidationError(f"细目表第 {row_index} 行核心考点不能为空")
        if score <= 0:
            raise ValidationError(f"细目表第 {row_index} 行分值必须大于 0")

        items.append(
            QuestionItem(
                number=number,
                question_type=question_type,
                core_concept=core_concept,
                score=score,
                difficulty=difficulty or "未标注",
                misconception=misconception or "未提供",
            )
        )

    if not items:
        raise ValidationError("细目表没有可用题目数据")
    return items


def create_score_template(items: Iterable[QuestionItem], output_path: str | Path, rows: int = 200) -> Path:
    questions = list(items)
    if not questions:
        raise ValidationError("没有题目数据，无法生成模板")

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "小题分"
    headers = [NAME_HEADER, TOTAL_HEADER] + [item.template_header for item in questions]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.freeze_panes = "C2"
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 12
    for col_idx, item in enumerate(questions, start=3):
        letter = sheet.cell(1, col_idx).column_letter
        sheet.column_dimensions[letter].width = min(max(len(item.template_header) + 4, 14), 24)
        validation = DataValidation(
            type="decimal",
            operator="between",
            formula1="0",
            formula2=str(item.score),
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="分数超出范围",
            error=f"本题分数必须在 0 到 {item.score:g} 之间",
        )
        sheet.add_data_validation(validation)
        validation.add(f"{letter}2:{letter}{rows + 1}")

    first_score_col = sheet.cell(1, 3).column_letter
    last_score_col = sheet.cell(1, len(headers)).column_letter
    for row_idx in range(2, rows + 2):
        sheet.cell(row_idx, 1).value = None
        sheet.cell(row_idx, 2).value = None

    sheet.auto_filter.ref = f"A1:{last_score_col}{rows + 1}"
    workbook.save(output)
    return output


def _parse_question_number_from_header(header: str, items_by_number: Dict[str, QuestionItem]) -> str | None:
    for number, item in items_by_number.items():
        if header == item.template_header:
            return number
    match = re.search(r"(\d+(?:\.\d+)?)（", header)
    if match and match.group(1) in items_by_number:
        return match.group(1)
    return None


def load_score_sheet(path: str | Path, items: Iterable[QuestionItem]) -> Tuple[List[StudentScores], List[str]]:
    questions = list(items)
    items_by_number = {item.number: item for item in questions}
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    headers = [_as_text(cell.value) for cell in sheet[1]]

    if len(headers) < 3 or headers[0] != NAME_HEADER or headers[1] != TOTAL_HEADER:
        raise ValidationError("小题分模板前两列必须是：姓名、全卷分数")

    column_to_number: Dict[int, str] = {}
    for col_idx, header in enumerate(headers[2:], start=3):
        if not header:
            continue
        number = _parse_question_number_from_header(header, items_by_number)
        if number:
            column_to_number[col_idx] = number

    missing = [item.template_header for item in questions if item.number not in set(column_to_number.values())]
    if missing:
        raise ValidationError(f"小题分模板缺少题目列：{', '.join(missing[:5])}")

    students: List[StudentScores] = []
    warnings: List[str] = []
    for row_idx in range(2, sheet.max_row + 1):
        name = _as_text(sheet.cell(row_idx, 1).value)
        if not name:
            if any(sheet.cell(row_idx, col).value not in (None, "") for col in range(2, sheet.max_column + 1)):
                warnings.append(f"第 {row_idx} 行没有姓名，已跳过")
            continue

        declared_total_raw = sheet.cell(row_idx, 2).value
        declared_total = None if declared_total_raw in (None, "") else _as_float(declared_total_raw, f"第 {row_idx} 行全卷分数")
        scores: Dict[str, float] = {}
        for col_idx, number in column_to_number.items():
            item = items_by_number[number]
            value = sheet.cell(row_idx, col_idx).value
            score = 0.0 if value in (None, "") else _as_float(value, f"第 {row_idx} 行第 {number} 题分数")
            if score < 0 or score > item.score:
                raise ValidationError(f"第 {row_idx} 行第 {number} 题分数 {score:g} 超出 0-{item.score:g}")
            scores[number] = score

        students.append(StudentScores(name=name, declared_total=declared_total, scores_by_question=scores, row_number=row_idx))

    if not students:
        raise ValidationError("小题分表中没有可分析的学生数据")
    return students, warnings
