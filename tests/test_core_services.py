from __future__ import annotations

import unittest
import shutil
from contextlib import contextmanager
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.analysis_service import analyze_batch
from app.exceptions import ValidationError
from app.excel_service import create_score_template, load_detail_table, load_score_sheet
from app.pdf_service import build_student_report


ROOT = Path(__file__).resolve().parents[1]
DETAIL_FILE = ROOT / "2026-庄市&甬江-七下期中测试卷-细目表.xlsx"
TMP_ROOT = ROOT / "test_outputs"
_TEMP_COUNTER = 0


@contextmanager
def temp_dir():
    global _TEMP_COUNTER
    _TEMP_COUNTER += 1
    path = TMP_ROOT / f"case_{_TEMP_COUNTER}"
    if path.exists():
        shutil.rmtree(path)
    path.mkdir(parents=True)
    try:
        yield str(path)
    finally:
        shutil.rmtree(path, ignore_errors=True)


class CoreServicesTest(unittest.TestCase):
    def test_load_detail_table_and_generate_template(self) -> None:
        items = load_detail_table(DETAIL_FILE)
        self.assertEqual(len(items), 33)
        self.assertEqual(items[0].template_header, "选择题1（2.0分）")

        with temp_dir() as tmp:
            output = create_score_template(items, Path(tmp) / "template.xlsx", rows=3)
            workbook = load_workbook(output, data_only=False)
            sheet = workbook.active
            self.assertEqual(sheet["A1"].value, "姓名")
            self.assertEqual(sheet["B1"].value, "全卷分数")
            self.assertEqual(sheet["C1"].value, "选择题1（2.0分）")
            self.assertTrue(str(sheet["B2"].value).startswith("=SUM("))

    def test_score_sheet_validation_rejects_over_max(self) -> None:
        items = load_detail_table(DETAIL_FILE)
        with temp_dir() as tmp:
            template = create_score_template(items, Path(tmp) / "template.xlsx", rows=2)
            workbook = load_workbook(template)
            sheet = workbook.active
            sheet["A2"] = "测试学生"
            sheet["C2"] = 99
            workbook.save(template)

            with self.assertRaises(ValidationError):
                load_score_sheet(template, items)

    def test_analyze_and_build_pdf(self) -> None:
        items = load_detail_table(DETAIL_FILE)
        with temp_dir() as tmp:
            template = create_score_template(items, Path(tmp) / "template.xlsx", rows=2)
            workbook = load_workbook(template)
            sheet = workbook.active
            sheet["A2"] = "满分学生"
            sheet["A3"] = "薄弱学生"
            for col_idx, item in enumerate(items, start=3):
                sheet.cell(2, col_idx).value = item.score
                sheet.cell(3, col_idx).value = 0 if col_idx % 2 == 0 else item.score
            workbook.save(template)

            students, warnings = load_score_sheet(template, items)
            self.assertEqual(warnings, [])
            analyses = analyze_batch(items, students)
            self.assertEqual(len(analyses), 2)
            self.assertGreater(analyses[0].score_rate, analyses[1].score_rate)

            pdf = build_student_report(analyses[1], Path(tmp) / "report.pdf")
            self.assertTrue(pdf.exists())
            self.assertGreater(pdf.stat().st_size, 1000)

    def test_missing_header_rejected(self) -> None:
        with temp_dir() as tmp:
            bad = Path(tmp) / "bad.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["题号", "题型"])
            workbook.save(bad)

            with self.assertRaises(ValidationError):
                load_detail_table(bad)


if __name__ == "__main__":
    unittest.main()
